import unittest
import tempfile
import os
import logging
from unittest.mock import patch
from auxiclean import Selector
from collections import OrderedDict
from openpyxl import Workbook


class TestBase(unittest.TestCase):
    # Cours (nom),Cours (code),Dispo,Programme
    courses = {}
    # Nom,Premier Choix,Deuxieme Choix,Troisieme Choix,Quatrieme Choix,
    # Cinquieme Choix,Disponibilite,Tp_total,Schoplarite,Nobel,
    # Programme d'etude,Cote Z
    candidates = {}
    ordered_candidates = None
    loglevel = logging.DEBUG

    def setUp(self):
        self.selector = None
        # use temporary file to do the tests
        self.tempdir = tempfile.TemporaryDirectory()
        self.data_path = os.path.join(self.tempdir.name,
                                      "auxiclean_unittests.xlsx")
        wb = Workbook()
        # save courses
        # data organized as follows:
        # code, name, disponibilities, discipline
        ws = wb.create_sheet("Affichage")
        ncourses = len(self.courses)
        title_cells = ("A1", "B1", "C1", "D1")
        titles = ("sigle", "titre du cours", "nombre de postes", "discipline")
        for tc, t in zip(title_cells, titles):
            ws[tc].value = t
        rows = ws.iter_rows(min_col=1, max_col=4, min_row=2,
                            max_row=ncourses + 1)
        for row, (course, course_data) in zip(rows,
                                              self.courses.items()):
            row[0].value = course_data["code"]
            row[1].value = course
            row[2].value = course_data["disponibilities"]
            row[3].value = course_data["discipline"]

        # save candidates
        # data organized as follows
        # name, max, scolarity, courses done, nobels, discipline, choices, gpa
        ws = wb.create_sheet("Candidatures")
        title_cells = ("A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1")
        titles = ("nom", "maximum", "cycle", "cours donnés", "nobels",
                  "discipline", "choix", "cote z")
        for tc, t in zip(title_cells, titles):
            ws[tc].value = t
        # build ordered candidates dict if necessary
        if self.ordered_candidates is not None:
            self.candidates = OrderedDict()
            for pair in self.ordered_candidates:
                self.candidates[pair[0]] = pair[1]
        ncandidates = len(self.candidates)
        rows = ws.iter_rows(min_col=1, max_col=8, min_row=2,
                            max_row=ncandidates + 1)
        for row, (c, c_data) in zip(rows, self.candidates.items()):
                row[0].value = c
                row[1].value = c_data["maximum"]
                row[2].value = c_data["scolarity"]
                row[3].value = ", ".join(c_data["courses given"])
                row[4].value = c_data["nobels"]
                row[5].value = c_data["discipline"]
                row[6].value = ", ".join(c_data["choices"])
                row[7].value = c_data["gpa"]
        wb.save(self.data_path)

    def tearDown(self):
        self.tempdir.cleanup()
        del self.tempdir
        del self.selector


class TestSelector(TestBase):
    # two different courses
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "générale"}}
    # two candidates each applying for a different course. No conflict
    candidates = {"Albert A": {"maximum": 2,
                               "scolarity": 2,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710", "1620"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["1441", ],
                               "gpa": 2.6},
                  "Claude C": {"maximum": 2,
                               "scolarity": 3,
                               "courses given": ["1651", "3131"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["2710", ],
                               "gpa": 3.0}}

    def test_running(self):
        # simple test that checks that both candidates receive
        # their first choice.
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Albert A")
        self.assertEqual(dist["2710"][0].name, "Claude C")
        # check that there is no other candidates in the distribution
        self.assertEqual(len(dist["1441"]), 1)
        self.assertEqual(len(dist["2710"]), 1)


class TestSortingSelector(TestBase):
    # One course with one place
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # two candidates apply for the same course but Alice gave it one time
    # So Alice has more experience than Bob, she should win the position.
    candidates = {"Bob": {"choices": ["1441", ],
                          "maximum": 1,
                          "scolarity": 2,
                          "courses given": [],
                          "discipline": "générale",
                          "nobels": 0,
                          "gpa": 3.6},
                  "Alice": {"choices": ["1441", ],
                            "maximum": 1,
                            "scolarity": 2,
                            "nobels": 0,
                            "courses given": ["1441", ],
                            "gpa": 3.0,
                            "discipline": "astro"}}

    def test_sort_by_specific_tp_experience(self):
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Alice")


class TestMultiplePosition(TestBase):
    # two different courses for one person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astro"}}
    # one candidate applying for two different course.
    # if the candidate is the best suited, he gets both choice
    # assuming he has two disponibilities
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 2,
                               "scolarity": 2,
                               "courses given": [],
                               "gpa": 3.0,
                               "nobels": 0,
                               "discipline": "générale"}}

    def test_two_class_for_one_person(self):
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Albert A")
        self.assertEqual(dist["2710"][0].name, "Albert A")


class TestSecondChoiceIsBetter(TestBase):
    # two different courses for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astrophysique"}}
    # both candidate have different first choice but one candidate
    # has the other class in his second choice and is better so
    # once his first choice hi passed, he also gets his second choice
    # assuming he has two disponibilities
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 2,
                               "scolarity": 2,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710"],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["2710", ],
                               "maximum": 1,
                               "scolarity": 3,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["2710", "2710"],
                               "discipline": "astrophysique"}}

    def test_second_choice_beat_first_if_better(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Albert A")
        self.assertEqual(dist["2710"][0].name, "Albert A")


class TestSecondChoiceIsBetterButNoMoreDispo(TestBase):
    # two different courses for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astrophysique"}}
    # both candidate have different first choice but one candidate
    # has the other class in his second choice and is better but
    # no more disponibilities so the other students still gets it
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710"],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["2710", ],
                               "maximum": 1,
                               "scolarity": 3,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["2710", "2710"],
                               "discipline": "astrophysique"}}

    def test_second_choice_beat_first_if_better_and_has_dispo(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Albert A")
        self.assertEqual(dist["2710"][0].name, "Claude C")


class TestNoDispo(TestBase):
    # two different courses for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astrophysique"}}
    # No more dispo for both students so nobody gets chosen
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 0,
                               "scolarity": 2,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710"],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["2710", ],
                               "maximum": 0,
                               "scolarity": 3,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["2710", "2710"],
                               "discipline": "astrophysique"}}

    def test_no_more_dispo(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"], [])
        self.assertEqual(dist["2710"], [])


class TestNoSpaceInClass(TestBase):
    # two different courses for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 0,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 0,
                         "discipline": "astrophysique"}}
    # no space in class so nobody gets chosen (for second showing)
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710"],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["2710", ],
                               "maximum": 1,
                               "scolarity": 3,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["2710", "2710"],
                               "discipline": "astrophysique"}}

    def test_no_more_space(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"], [])
        self.assertEqual(dist["2710"], [])


class TestSwitch(TestBase):
    # two different courses for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"},
               "Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astrophysique"}}
    # Second choice of 1 candidate is better than another one's first
    # So they switch places
    candidates = {"Albert A": {"choices": ["1441", "2710"],
                               "maximum": 2,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710"],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["2710", "1441"],
                               "maximum": 1,
                               "scolarity": 3,
                               "gpa": 3.0,
                               "nobels": 0,
                               "courses given": ["2710", "2710",
                                                 "1441", "1441"],
                               "discipline": "particules"}}

    def test_switch(self):
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Claude C")
        self.assertEqual(dist["2710"][0].name, "Albert A")


class TestCourseGiven(TestBase):
    # one course for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # If number of tp specific are equal, we look at total tp
    candidates = {"Albert A": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", ],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "2710", ],
                               "discipline": "particules"}}

    def test_course_given(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Claude C")


class TestScholarity(TestBase):
    # one course for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # If number of total tp are equal, we look at scholarity
    candidates = {"Albert A": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "1652", ],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 3,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "2710", ],
                               "discipline": "particules"}}

    def test_scholarity(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Claude C")


class TestNobel(TestBase):
    # one course for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # If scholarity is equal, we look at number of nobels won
    candidates = {"Albert A": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "1652", ],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 3,
                               "courses given": ["1441", "2710", ],
                               "discipline": "particules"}}

    def test_nobel(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Claude C")


class TestProgram(TestBase):
    # one course for two person
    courses = {"Astro": {"code": "2710",
                         "disponibilities": 1,
                         "discipline": "astrophysique"}}
    # If number of nobels won are equal, we look at program of research
    candidates = {"Albert A": {"choices": ["2710", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "1652", ],
                               "discipline": "particules"},
                  "Claude C": {"choices": ["2710", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "2710", ],
                               "discipline": "astrophysique"}}

    def test_program(self):
        self.selector = Selector(self.data_path)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["2710"][0].name, "Claude C")


class TestGPA(TestBase):
    # one course for two person
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # If both candidates are studying in related program,
    # we look at their GPA
    candidates = {"Albert A": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 2.0,
                               "nobels": 2,
                               "courses given": ["1441", "1652", ],
                               "discipline": "astrophysique"},
                  "Claude C": {"choices": ["1441", ],
                               "maximum": 1,
                               "scolarity": 2,
                               "gpa": 3.0,
                               "nobels": 2,
                               "courses given": ["1441", "2710", ],
                               "discipline": "particules"}}

    def test_gpa(self):
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        # results
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Claude C")


@patch('auxiclean.user_input.get_user_input')
class TestUserInput(TestBase):
    # one course, two equal candidates
    courses = {"Electro": {"code": "1441",
                           "disponibilities": 1,
                           "discipline": "générale"}}
    # ordered dict here because for python v < 3.6, dict order is not
    # guaranteed and we want to assign albert A as choice #1
    ordered_candidates = [("Albert A", {"choices": ["1441", ],
                                        "maximum": 1,
                                        "scolarity": 2,
                                        "gpa": 2.0,
                                        "nobels": 2,
                                        "courses given": ["1441", "1652", ],
                                        "discipline": "astrophysique"}),
                          ("Bernard B", {"choices": ["1441", ],
                                         "maximum": 1,
                                         "scolarity": 2,
                                         "gpa": 2.0,
                                         "nobels": 2,
                                         "courses given": ["1441",
                                                           "2710", ],
                                         "discipline": "particules"})]

    def test_user_input_simple(self, user_input_mock):
        # test that user input chooses first candidate over second.
        user_input_mock.side_effect = ["1", "oui"]
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Albert A")

    def test_user_input_NaN(self, user_input_mock):
        # test that the code still selects the good candidates
        # if user do not enter a number the first time
        user_input_mock.side_effect = ["not a number", "2", "oui"]
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Bernard B")

    def test_user_input_less_than_1(self, user_input_mock):
        # test that the code still selects the good candidates
        # if user enters a number less than 1
        user_input_mock.side_effect = ["0", "2", "oui"]
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Bernard B")

    def test_user_input_retry(self, user_input_mock):
        # test that the code still selects the good candidates
        # if user enters 'no' at the last step to retry selection
        # (test when user changes its mind)
        user_input_mock.side_effect = ["1", "no", "2", "yes"]
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        second_name = self.selector.excel_mgr.candidates[1].name
        self.assertEqual(dist["1441"][0].name, second_name)

    def test_user_input_wrong_retry(self, user_input_mock):
        # test that the code still selects the good candidates
        # if user enters neither 'yes' or 'no' at the
        # last step to retry selection
        # (test when user changes its mind but makes a typo)
        user_input_mock.side_effect = ["1", "not yes or no", "N", "2", "y"]
        self.selector = Selector(self.data_path, loglevel=self.loglevel)
        dist = self.selector.distribution
        self.assertEqual(dist["1441"][0].name, "Bernard B")
