from .test_selector import TestBase
from openpyxl import load_workbook
from auxiclean import Selector
from collections import OrderedDict
import warnings


class TestExcelManager(TestBase):
    # two different courses
    courses = OrderedDict({"Electro": {"code": "1441",
                                       "disponibilities": 1,
                                       "discipline": "générale"},
                           "Astro": {"code": "2710",
                                     "disponibilities": 1,
                                     "discipline": "générale"}})
    # two candidates each applying for a different course. No conflict
    candidates = {"Albert A": {"maximum": 2,
                               "scolarity": 2,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710", "1620"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["1441", ],
                               "gpa": 2.6},
                  "Claude C": {"maximum": 2,
                               "scolarity": 3,
                               "courses given": ["1651", "3131"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["2710", ],
                               "gpa": 3.0}}

    def test_writing_distribution(self):
        # make distribution
        self.selector = Selector(self.data_path)
        # open excel file
        wb = load_workbook(self.data_path)
        # check that the sheet exists
        self.assertIn("Distribution", wb.sheetnames)
        # get sheet
        ws = wb["Distribution"]
        # check that first row is the titles
        self.assertEqual(ws["A1"].value.lower(), "sigle")
        self.assertEqual(ws["B1"].value.lower(), "distribution")
        # check that next rows are courses with selection
        allcodes = [x["code"] for x in self.courses.values()]
        self.assertIn(ws["A2"].value, allcodes)
        self.assertIn(ws["A3"].value, allcodes)
        # only two courses, next rows should be empty
        self.assertIs(ws["A4"].value, None)
        for row in ws.iter_rows(min_col=1, max_col=2, min_row=2, max_row=3):
            if row[0].value == "1441":
                self.assertEqual(row[1].value, "Albert A")
            else:
                self.assertEqual(row[1].value, "Claude C")

    def test_distribution_sheet_already_exists(self):
        # load workbook and add a distribution cheet
        wb = load_workbook(self.data_path)
        # create distribution sheet
        ws = wb.create_sheet("Distribution")
        safe_string = "Check that this string is not erased."
        ws["A1"].value = safe_string
        wb.save(self.data_path)
        del wb
        # call selector and check that already existing sheet is not erased
        # also check that a user warning is raised
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")
            self.selector = Selector(self.data_path)
            # check a warning is raised (only one)
            self.assertEqual(len(w), 1)
            self.assertIn("Distribution sheet already exists",
                          str(w[-1].message))
        # reload wb
        wb = load_workbook(self.data_path)
        self.assertIn("Distribution1", wb.sheetnames)
        # check that the previous sheet was not overridden
        ws = wb["Distribution"]
        self.assertEqual(ws["A1"].value, safe_string)
        del wb


class TestExcelCandidateChoiceError(TestBase):
    # two different courses
    courses = OrderedDict({"Electro": {"code": "1441",
                                       "disponibilities": 1,
                                       "discipline": "générale"},
                           "Astro": {"code": "2710",
                                     "disponibilities": 1,
                                     "discipline": "générale"}})
    # two candidates each applying for a different course. No conflict
    # one of a candidate's choice not in the course list.
    candidates = {"Albert A": {"maximum": 2,
                               "scolarity": 2,
                               "courses given": ["1441", "2710", "2710",
                                                 "2710", "2710", "1620"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["1441", "1234"],  # not in courses
                               "gpa": 2.6},
                  "Claude C": {"maximum": 2,
                               "scolarity": 3,
                               "courses given": ["1651", "3131"],
                               "nobels": 0,
                               "discipline": "générale",
                               "choices": ["2710", ],
                               "gpa": 3.0}}

    def test_raise_choice_error(self):
        with self.assertRaises(ValueError):
            self.selector = Selector(self.data_path)
