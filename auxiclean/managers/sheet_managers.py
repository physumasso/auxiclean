from ..candidate import Candidate
from ..course import Course
from ..exceptions import ExcelError
from openpyxl.styles import Alignment, colors, PatternFill, Font
import logging


class BaseSheetManager:
    _loggername = None

    def __init__(self, workbook, loglevel=logging.INFO):
        self.wb = workbook
        logging.basicConfig()
        self._logger = logging.getLogger(self._loggername)
        self._logger.setLevel(loglevel)

    def _find_max_columns(self, titles_row):
        for i, cell in enumerate(titles_row):
            if cell.value is None:
                return i
            elif cell.value.lower() == "none":
                return i
        return len(titles_row)

    def get_sheet(self, workbook, sheetname):
        for sheet in workbook.worksheets:
            if sheet.title.lower() == sheetname.lower():
                return sheet
        raise ExcelError("%s sheet is not found inside file." % sheetname)


class CoursesSheetManager(BaseSheetManager):
    _loggername = "auxiclean.managers.courses_sheet_manager"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._courses = None

    @property
    def courses(self):
        if self._courses is None:
            self._courses = self.load_courses_from_wb(self.wb)
        return self._courses

    def load_courses_from_wb(self, wb):
        # get the worksheet
        ws = self.get_sheet(wb, "Affichage")
        titles = ws[1]  # assume columns are labeled in first row
        max_column = self._find_max_columns(titles)
        titles = titles[:max_column]
        courses_data = ws[2:ws.max_row]  # rest are actual courses data
        if ws.max_row == 2:
            courses_data = (courses_data, )  # bring into a tuple if needed
        courses_data = [c[:max_column] for c in courses_data]
        courses = []
        for course in courses_data:
            d = {label.value.lower(): cell.value
                 for label, cell in zip(titles, course)}
            courses.append(d)
        return [Course(x["titre du cours"],
                       x["sigle"],
                       x["nombre de postes"],
                       x["discipline"]) for x in courses]


class CandidatesSheetManager(BaseSheetManager):
    _loggername = "auxiclean.managers.candidate_sheet_manager"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._candidates = None

    @property
    def candidates(self):
        if self._candidates is None:
            self._candidates = self.load_candidates_from_wb(self.wb)
        return self._candidates

    def load_candidates_from_wb(self, wb):
        ws = self.get_sheet(wb, "Candidatures")
        titles = ws[1]
        max_column = self._find_max_columns(titles)
        titles = titles[:max_column]
        candidates_data = ws[2:ws.max_row]
        if ws.max_row == 2:
            candidates_data = (candidates_data, )  # bring into a tuple
        candidates_data = [c[:max_column] for c in candidates_data]
        candidates = []
        for i, candidate in enumerate(candidates_data):
            d = {label.value.lower(): cell.value
                 for label, cell in zip(titles, candidate)}
            d["choix"] = self._get_list_from_str(d["choix"])
            d["cours donnés"] = self._get_list_from_str(d["cours donnés"])
            self._checkup_candidate_data(d, i)
            candidates.append(d)
        return [Candidate(x["nom"],
                          x["maximum"],
                          x["cours donnés"],
                          x["cycle"],
                          x["nobels"],
                          x["discipline"],
                          x["cote z"], *x["choix"]) for x in candidates]

    def _checkup_candidate_data(self, data, i):
        # i == candidate index row
        if data["nom"] is None:
            raise ExcelError("Candidature ligne %i n'a pas de nom." % i)
        if data["maximum"] is None:
            raise ExcelError("Candidature %s n'a pas de maximum." %
                             data["nom"])
        if data["cycle"] is None:
            raise ExcelError("Candidature %s n'a pas de cycle d'étude." %
                             data["nom"])
        if data["choix"] is None:
            raise ExcelError("Candidature %s n'a pas de choix!" % data["nom"])

    def _get_list_from_str(self, string):
        # string format: 1441-90, 3131, 2810, ...
        # courses given in order of preferences, separated with a comma
        if string is None:
            return []
        split = string.split(",")
        return [x.strip() for x in split]


class DistributionSheetManager(BaseSheetManager):
    _loggername = "auxiclean.managers.distribution_sheet_manager"

    def write_distribution(self, distribution):
        # get sheet
        ws = self.get_sheet(self.wb)
        # write headers
        self.write_distribution_headers(ws)
        # write distribution
        self._do_write_distribution(ws, distribution)

    def _do_write_distribution(self, worksheet, distribution):
        nrows_required = len(distribution.keys())
        rows = worksheet.iter_rows(min_col=1, max_col=2, min_row=2,
                                   max_row=nrows_required + 1)
        font = Font(name="Arial", bold=False, color=colors.BLACK, size=10)
        alignment = Alignment(horizontal="left", vertical="center")
        for row, (course, selection) in zip(rows, distribution.items()):
            # course name in first cell
            cell_course = row[0]
            cell_course.font = font
            cell_course.alignment = alignment
            cell_course.value = course
            # course distribution in second cell
            cell_selection = row[1]
            cell_selection.font = font
            cell_selection.alignment = alignment
            cell_selection.value = ", ".join([x.name for x in selection])

    def write_distribution_headers(self, worksheet):
        # header = course code, distribution
        headers = ["Sigle", "Distribution"]
        header_cells = worksheet.iter_rows(min_col=1, max_col=2,
                                           min_row=1, max_row=1)
        # fill cell with a black background
        fill = PatternFill(start_color=colors.BLACK,
                           end_color=colors.BLACK,
                           fill_type="solid")
        font = Font(name="Arial",
                    bold=True,
                    size=11,
                    color=colors.WHITE)
        alignment = Alignment(horizontal="left", vertical="center")
        for row in header_cells:
            for header, cell in zip(headers, row):
                cell.value = header
                cell.font = font
                cell.fill = fill
                cell.alignment = alignment

    def get_sheet(self, workbook):
        # get distribution sheet
        # it one already exists, create one but warn user.
        try:
            ws = super().get_sheet(workbook, "Distribution")
        except ExcelError:
            # sheet does not exist
            ws = workbook.create_sheet("Distribution")
        else:
            # sheet exists, warn user, but create one
            ws = workbook.create_sheet("Distribution")
            self._logger.warning("A Distribution sheet already exists"
                                 " in destination."
                                 " A new one will be created: %s." % ws.title)
        return ws
